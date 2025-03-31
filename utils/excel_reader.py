import openpyxl

def read_test_data_from_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    test_data = [(sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value) for i in range(2, sheet.max_row + 1)]
    return test_data
